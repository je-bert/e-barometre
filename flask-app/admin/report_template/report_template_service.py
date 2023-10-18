from flask import render_template, send_file, request, make_response, jsonify, render_template_string
import os
from openpyxl import load_workbook
from api.results import output_from_file

def get_view():
  return render_template('report-template.html')

def export_report_template():
  file_name = 'master-results-template.xlsx'
  # confirm file exists
  if os.path.exists(file_name):
    return send_file(file_name, as_attachment=True)
  else:
    # return not found
    return make_response(jsonify({"message": "File not found"}), 404)

def import_report_template():
  file = request.files['file']
  file_name = 'master-results-template.xlsx'
  file.save(file_name)
  wb = load_workbook(file_name)
  ws = wb.get_sheet_by_name("TEST_pour PROTOTYPE")
  for row in range(4, 411):
    ws.cell(row=row, column=4).value = ""
  wb.save(file_name)
  return make_response(jsonify({"message": "File uploaded"}), 200)

def test_remport_template():
  file = request.files['file']
  file_name = 'test-report-template.xlsx'
  file.save(file_name)
  html = output_from_file('test-report-template.xlsx')

  # Clean up
  if os.path.exists(file_name):
    os.remove(file_name)

  return render_template_string(html)