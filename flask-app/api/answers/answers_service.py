
from flask import request, jsonify, abort
from models import User, Question, Answer
from database import db
from datetime import datetime
from shutil import copyfile
import os
from openpyxl import load_workbook

def create(current_user):
  data = request.json
  user_id = current_user.user_id
  question_id = data['question_id']
  value= data['value']
  date_created = datetime.now()
  answer = Answer(user_id=user_id,value=value, question_id=question_id, date_created=date_created)
  db.session.add(answer)
  db.session.commit()
  return jsonify(answer), 200



def update(current_user):

  user = User.query\
    .filter_by(user_id = current_user.user_id)\
    .first()

  if not user:
      return abort(400)

  if not os.path.exists('master_results'):
    os.makedirs('master_results')

  template_file = 'master_results_template.xlsx'
  output_file = 'master_results/{}.xlsx'.format(user.user_id)
  if not os.path.exists(output_file):
    copyfile(template_file, output_file)

  # Load the workbook and access the sheet we'll paste into
  wb = load_workbook(output_file)

  # Demo code
  ws = wb.worksheets[0]

  json = request.json

  if not 'answers' in json:
    return "Erreur", 400

  for answer in json['answers']:
    if not 'question_id' in answer:
      return "Erreur", 400

    if not 'value' in answer:
      return "Erreur", 400

    question = Question.query\
        .filter_by(question_id = answer['question_id'])\
        .first()
    
    if not question:
      return "Erreur", 400
    
    #TODO: Validation answer + custom answer

    new_answer = Answer(question_id = answer['question_id'], user_id = current_user.user_id, value = answer['value'], date_created = datetime.now())
    db.session.merge(new_answer)

    if question.type == 'select-multiple':
      answers = answer['value'].split(',')
      i = 0
      for cell in ws["C"]:
        if type(cell).__name__ != 'MergedCell' and cell.internal_value == question.question_id:
          ws.cell(row=cell.row, column=4).value = 1 if str(i) in answers else 0
          i += 1
    else:
      for cell in ws["C"]:
        if type(cell).__name__ != 'MergedCell' and cell.internal_value == question.question_id:
            ws.cell(row=cell.row, column=4).value = answer['value']

  db.session.commit()
  wb.save(output_file)
    
  return jsonify("Vos réponses ont été sauvegardées!"), 200