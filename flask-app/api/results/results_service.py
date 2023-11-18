from flask import abort, jsonify, render_template
from models import User, Answer, Question, AnalysisSection, AnalysisSubsection, Report, Invoice
import os
from database import db
from shutil import copyfile
import os
from openpyxl import load_workbook
from pycel import ExcelCompiler
import os
from pyhtml2pdf import converter
from api.invoices.invoices_service import get_user_subscription



TEMPLATE_FILE = 'master-results-template.xlsx'

def generate(user_id):
  subscription = get_user_subscription(user_id)

  if subscription == None:
    return jsonify({"message":"Vous n'avez pas de souscription active."}), 400
  
  user = User.query\
    .filter_by(user_id = user_id)\
    .first()

  if not user:
      return abort(400)
  
  if not os.path.exists('master_results'):
    os.makedirs('master_results')

  template_file = TEMPLATE_FILE
  worksheet_name = 'TEST_pour PROTOTYPE'

  report = Report.query\
    .filter_by(user_id = user_id, is_completed = False)\
    .first()
  
  if report == None:
    return jsonify({"message":"Aucun rapport en cours."}), 400

  db.session.add(report)
  db.session.commit()

  output_file = 'master_results/{}.xlsx'.format(report.report_id)
  
  if os.path.exists(output_file):
    os.remove(output_file)

  if not os.path.exists(output_file):
     copyfile(template_file, output_file)

  # Load the workbook and access the sheet we'll paste into
  wb = load_workbook(output_file)
  ws = wb.get_sheet_by_name(worksheet_name)
  excel = ExcelCompiler(filename=output_file)



  answers = Answer.query\
    .filter_by(report_id = report.report_id)\
    .all()

  for answer in answers:
    if answer.value == '-1':
      continue
    question = Question.query\
          .filter_by(question_id = answer.question_id )\
          .first()
    if question.type == 'select-multiple':
      values = answer.value.split(',')
      i = 1
      for cell in ws["C"]:
        value = excel.evaluate(f"'{worksheet_name}'!{cell.coordinate}")
        if value == answer.question_id:
          ws.cell(row=cell.row, column=4).value = (1 if str(i) in values else 0)
          i += 1
    else:
      for cell in ws["C"]:
        value = excel.evaluate(f"'{worksheet_name}'!{cell.coordinate}")
        if value == answer.question_id:
            ws.cell(row=cell.row, column=4).value = answer.value
  
  wb.save(output_file)
 
  #Creation of pdf
  html_content = output(report.report_id)

  with open(f'master_results/{report.report_id}.html', 'w',encoding='utf-8') as f:
      f.write(html_content)

  source = os.path.abspath(f'master_results/{report.report_id}.html')
  target = os.path.abspath(f'master_results/{report.report_id}.pdf')
  converter.convert(f'file:///{source}', target, print_options={'marginTop': 0, 'marginRight': 0, 'marginBottom': 0, 'marginLeft': 0})

  report.is_completed = True
  # expire unique invoice
  if subscription != 'multiple':
    Invoice.query\
      .filter_by(user_id = user_id, status = 'paid', product_id = subscription)\
      .update({Invoice.status: 'expired', Invoice.date_expiration: db.func.current_timestamp()})
  db.session.commit()

  return jsonify("Votre rapport a été généré!"), 200

def output_from_file(file_name):

  if not os.path.exists(file_name):
    return abort(404)
  worksheet_name = 'Test de contenu du rapport'

  # sections = db.session.query(AnalysisSection)\
  #   .join(AnalysisSubsection)\
  #   .order_by(AnalysisSection.order)\
  #   .all()

  wb = load_workbook(file_name)

  # wb['Test de contenu du rapport']['A411'].value = 'allo'

  # wb.save(file_name)


  excel = ExcelCompiler(filename=file_name)

  sections = []
  themes = []
  analysis = []
  flag_introduction = None
  yellow_flags = []
  red_flags = []
  previous_cell = None
  about = None
  insert_page_break = False



  for i in range (1, 400):
    cell = excel.evaluate(f"'{worksheet_name}'!C{i}")
    if cell and cell != '':
      value = excel.evaluate(f"'{worksheet_name}'!E{i}")
      if previous_cell == 'red_flag' and cell != 'red_flag':
        sections.append(render_template('reports/themes.html', analysis_items = analysis, observations = themes))
        if len(red_flags) > 0 or len(yellow_flags) > 0:
          sections.append(render_template('reports/flags.html', yellow_flags = yellow_flags, red_flags = red_flags, flag_introduction = flag_introduction))
        themes = []
        analysis = []
        yellow_flags = []
        red_flags = []
        flag_introduction = None
        about = None
      previous_cell = cell
      if not value or value == '' or value == ' ' or value == 0 or value == '0':
        continue
      
      elif cell == 'section-title':
        if insert_page_break == True:
          insert_page_break = False
          sections.append('<div class="insert-page-break"></div>')
        sections.append(render_template('reports/section-title.html', content = value))
      elif cell == 'subsection-title':
        if insert_page_break == False:
          insert_page_break = True
        else:
          sections.append('<div class="insert-page-break"></div>')
        sections.append(render_template('reports/subsection-title.html', content = value))
      elif cell == 'about-barometer':
        sections.append(render_template('reports/about-barometer.html', content = value))
        about = value
      elif cell == 'barometer':
        if value == 'barometer-1' or value == 'barometer-5' or value == 'barometer-6':
          sections.pop()
          sections.append(render_template('reports/about-barometer.html', content = about, hide_lg = True))
        sections.append(render_template("reports/report-1/" + value + ".html", data = generate_barometer_data(value, excel, about), about = about))
      elif cell == 'ressources':
        sections.append(render_template('reports/ressources.html', content = value))
      elif cell == 'theme':
        themes.append(value)
      elif cell == 'analysis':
        analysis.append(value)
      elif cell == 'flag_introduction':
        flag_introduction = value
      elif cell == 'yellow_flag':
        yellow_flags.append(value)
      elif cell == 'red_flag':
        red_flags.append(value)
      print("Cell " + str(i))
      print("From workbook")
      print(wb['Test de contenu du rapport']['B' + str(i)].value)
      print("From pycell")
      print(cell)
      print(value)

  return render_template('reports/report-1/base.html', children = sections)

def output(report_id):
  if not os.path.exists('master_results'):
    os.makedirs('master_results')

  file_name = 'master_results/{}.xlsx'.format(report_id)

  return output_from_file(file_name)
  
 

def convert_xlookup_to_index_match():
  filename = TEMPLATE_FILE
  sheetname = 'Modèle Calcul A-R + Sommaire'
  cell_range = 'D32:U800'

  # Load the workbook
  workbook = load_workbook(filename)
  worksheet = workbook[sheetname]

  # Get the cell range
  cells = worksheet[cell_range]

  for row in cells:
      for cell in row:
          # Get the formula from the cell
          formula = cell.value
          # Check if the formula is an XLOOKUP formula
          if formula and not isinstance(formula, int) and  formula.startswith('=_xlfn.XLOOKUP('):
              lookup_args = formula[15:-1].split(',')
              search_value = lookup_args[0].strip()
              lookup_range = lookup_args[1].strip()
              return_range = lookup_args[2].strip()

              # Replace the XLOOKUP formula with INDEX and MATCH formula
              index_match_formula = f'=_xlfn.INDEX({return_range}, _xlfn.MATCH({search_value}, {lookup_range}, 0), 1)'
              cell.value = index_match_formula

  # Save the updated workbook
  workbook.save(filename)
  return "Success",200

def generate_barometer_data(barometer, excel, about):
  if barometer == "barometer-1":
    data = {
      "id": "barometer_1",
      "value": excel.evaluate("'TEST_pour PROTOTYPE'!C438"),
      "range": [
        {
          "name": excel.evaluate("'TEST_pour PROTOTYPE'!G{}".format(i + 438)),
          "min": excel.evaluate("'TEST_pour PROTOTYPE'!H{}".format(i + 438)),
          "max": excel.evaluate("'TEST_pour PROTOTYPE'!I{}".format(i + 438)),
        } for i in range(4)
      ] ,
    }
    for r in data['range']:
      if r['min'] <= data['value'] <= r['max']:
        data['result'] = r['name']
        break
    data['value'] = (data['value'] / data['range'][3]['max']) * 100
    data['green'] = (data['range'][0]['max'] / data['range'][3]['max']) * 100
    data['yellow'] = (data['range'][1]['max'] / data['range'][3]['max']) * 100
    data['orange'] = (data['range'][2]['max'] / data['range'][3]['max']) * 100
    return data
  elif barometer == "barometer-2":
    data = {
      "id": "barometer_2",
      "label_1": excel.evaluate("'TEST_pour PROTOTYPE'!E453"),
      "label_2": excel.evaluate("'TEST_pour PROTOTYPE'!D453"),
      "items":[
        {
          "name": excel.evaluate("'TEST_pour PROTOTYPE'!C{}".format(i + 455)),
          "value_1": excel.evaluate("'TEST_pour PROTOTYPE'!E{}".format(i + 455)),
          "value_2": excel.evaluate("'TEST_pour PROTOTYPE'!D{}".format(i + 455))
        } for i in range(8) if excel.evaluate("'TEST_pour PROTOTYPE'!C{}".format(i + 455))
      ],
    }
    # Sort the items based on the difference between first_value and second_value
    data['items'] = sorted(data['items'], key=lambda x: abs(x['value_1'] - x['value_2']), reverse=True)
    return data
  elif barometer == "barometer-3":
    data = {
      "id": "barometer_3",
      "label_1": excel.evaluate("'TEST_pour PROTOTYPE'!E453"),
      "label_2": excel.evaluate("'TEST_pour PROTOTYPE'!D453"),
      "items":[
        {
          "name": excel.evaluate("'TEST_pour PROTOTYPE'!C{}".format(i + 464)),
          "value_1": excel.evaluate("'TEST_pour PROTOTYPE'!E{}".format(i + 464)),
          "value_2": excel.evaluate("'TEST_pour PROTOTYPE'!D{}".format(i + 464))
        } for i in range(6) if excel.evaluate("'TEST_pour PROTOTYPE'!C{}".format(i + 464))
      ],
    }
    # Sort the items based on the difference between first_value and second_value
    data['items'] = sorted(data['items'], key=lambda x: abs(x['value_1'] - x['value_2']), reverse=True)
    return data
  elif barometer == "barometer-4":
    data = {
      "id": "barometer_4",
      "items": [
        {
          "category": excel.evaluate("'TEST_pour PROTOTYPE'!N{}".format(485 + i)),
          "weight": excel.evaluate("'TEST_pour PROTOTYPE'!O{}".format(485 + i)),
          "behaviors": [],
        } for i in range(7)
      ],
    }
    for item in data['items']:
      for i in range(44):
        if excel.evaluate("'TEST_pour PROTOTYPE'!AM{}".format(470 + i)) == 'à inclure' and excel.evaluate("'TEST_pour PROTOTYPE'!AJ{}".format(470 + i)) == item['category']:
          behavior = {
            "name": excel.evaluate("'TEST_pour PROTOTYPE'!AK{}".format(470 + i)),
            "weight": excel.evaluate("'TEST_pour PROTOTYPE'!AL{}".format(470 + i)) 
          }
          item['behaviors'].append(behavior)
    # Sort behaviors of each item by their weight
    for item in data['items']:
      item['behaviors'].sort(key=lambda x: x['weight'], reverse=True)
    
    # Replace behaviors with behavior values without weight
    for item in data['items']:
      item['behaviors'] = [behavior['name'] for behavior in item['behaviors']]
    return data
  elif barometer == "barometer-5":
    data = {
      "id": "barometer_5",
      "value": excel.evaluate("'TEST_pour PROTOTYPE'!C524"),
      "range": [
        {
          "name": excel.evaluate("'TEST_pour PROTOTYPE'!F{}".format(i + 525)),
          "min": excel.evaluate("'TEST_pour PROTOTYPE'!G{}".format(i + 525)),
          "max": excel.evaluate("'TEST_pour PROTOTYPE'!H{}".format(i + 525)),
        } for i in range(4)
      ] ,
    }
    for r in data['range']:
      if r['min'] <= data['value'] <= r['max']:
        data['result'] = r['name']
        break
    data['value'] = (data['value'] / data['range'][3]['max']) * 100
    data['green'] = (data['range'][0]['max'] / data['range'][3]['max']) * 100
    data['yellow'] = (data['range'][1]['max'] / data['range'][3]['max']) * 100
    data['orange'] = (data['range'][2]['max'] / data['range'][3]['max']) * 100
    return data
  elif barometer == "barometer-6":
    data = {
      "id": "barometer_6",
      "value": excel.evaluate("'TEST_pour PROTOTYPE'!C545"),
      "range": [
        excel.evaluate("'TEST_pour PROTOTYPE'!D545"), # Green
        excel.evaluate("'TEST_pour PROTOTYPE'!E545"), # Yellow (min for risk)
        excel.evaluate("'TEST_pour PROTOTYPE'!F545"), # Orange
        excel.evaluate("'TEST_pour PROTOTYPE'!G545"), # Red
        excel.evaluate("'TEST_pour PROTOTYPE'!H545"), # Max
      ]
    }
    return data
  elif barometer == "barometer-7":
    data = {
      "id": "barometer_7",
      "items": [],
    }
   
    for i in range(30):
      value = excel.evaluate("'TEST_pour PROTOTYPE'!E{}".format(i + 571))
      if isinstance(value, int) and value > 0 and value:
        title = excel.evaluate("'TEST_pour PROTOTYPE'!D{}".format(i + 571))
        data['items'].append({"name": title, "value": value})
        # Maximum of 7 items
        if len(data['items']) >= 7:
          break
    data['items'] = sorted(data['items'], key=lambda d: d['value'], reverse=True)
    return data
  else:
    return {}















# je n'ai jamais aucun Constat pour réponse de l'enfant dans mes tests, la section TEST_pour PROTOTYPE!AI570 est systématiquement vide pour moi



# 'Cpts miroirs des parents' A27 -> if(and)
# 'Contribution NC' I -> tout mettre le bon lookup
